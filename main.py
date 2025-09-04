from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, Response
from pydantic import BaseModel
from typing import Dict, Optional
import uuid
import json
import io
from datetime import datetime
import numpy as np
import pandas as pd
import redis
import hashlib
import logging
import math
import base64
 

 

# PDF and Excel imports
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.barcharts import VerticalBarChart
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import PieChart, Reference

# FastAPI app
app = FastAPI(title="Taxync API", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Allow all origins for development
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Basic logging setup
logging.basicConfig(level=logging.INFO)

# Redis setup
try:
    redis_client = redis.Redis(host='localhost', port=6379, db=0, decode_responses=True)
    redis_client.ping()
except:
    redis_client = None

# Models
class TaxCalculationRequest(BaseModel):
    income: float
    section80C: float = 0
    section80D: float = 0
    hra: float = 0
    home_loan_interest: float = 0
    standard_deduction: float = 50000
    edu_loan_interest: float = 0
    donations: float = 0

class ExportRequest(BaseModel):
    formData: dict
    taxResult: dict
    chartImage: Optional[str] = None

class ShareRequest(BaseModel):
    formData: dict
    taxResult: dict

class UserInfo(BaseModel):
    name: str
    email: str
    generatedAt: str

# --- Tax Calculation Logic ---
def get_tax_saving_for_investment(investment_amount: float, current_income: float, current_deductions: float) -> float:
    """Calculates the potential tax saving from an additional investment."""
    tax_before = calculate_old_regime_tax(current_income, current_deductions)
    tax_after = calculate_old_regime_tax(current_income, current_deductions + investment_amount)
    return max(0, tax_before - tax_after)

def get_tax_slabs(regime: str) -> List[Dict[str, float]]:
    # This function is not used anywhere in the code, so it's not refactored
    pass

def calculate_old_regime_tax(income: float, total_deductions: float = 0) -> float:
    """
    Calculates tax based on the old regime slabs for FY 2025-26.
    Deductions are passed as a single total amount.
    """
    total_deductions = sum(deductions.values()) if deductions else 0
    taxable_income = max(0, income - total_deductions)

    # Section 87A rebate: If taxable income <= ₹5,00,000, tax liability is zero (before cess)
    if taxable_income <= 500000:
        return 0.0

    # FY 2025-26 Old Regime slabs
    slabs = np.array([0, 250000, 500000, 1000000])
    rates = np.array([0, 0.05, 0.20, 0.30])

    tax = 0.0
    for i in range(len(slabs)):
        if i == len(slabs) - 1:
            if taxable_income > slabs[i]:
                tax += (taxable_income - slabs[i]) * rates[i]
        else:
            if taxable_income > slabs[i]:
                slab_income = min(taxable_income, slabs[i + 1]) - slabs[i]
                tax += slab_income * rates[i]

    # 4% cess and rounding
    tax_with_cess = tax * 1.04
    return float(round(tax_with_cess))

def calculate_new_regime_tax(income: float, standard_deduction: float = 50000) -> float:
    """Calculate tax under New Regime for FY 2025-26.

    - Applies standard deduction only (default ₹50,000).
    - Slabs: 0–3L:0%, 3–6L:5%, 6–9L:10%, 9–12L:15%, 12–15L:20%, 15L+:30%
    - Adds 4% Health & Education cess at the end.
    - Rounds final tax to nearest integer.
    """
    taxable_income = max(0.0, float(income) - float(standard_deduction))

    # FY 2025-26 New Regime slabs
    slabs = np.array([0, 300000, 600000, 900000, 1200000, 1500000])
    rates = np.array([0, 0.05, 0.10, 0.15, 0.20, 0.30])

    tax = 0.0
    for i in range(len(slabs)):
        if i == len(slabs) - 1:
            if taxable_income > slabs[i]:
                tax += (taxable_income - slabs[i]) * rates[i]
        else:
            if taxable_income > slabs[i]:
                slab_income = min(taxable_income, slabs[i + 1]) - slabs[i]
                tax += slab_income * rates[i]

    tax_with_cess = tax * 1.04
    return float(round(tax_with_cess))

def generate_tax_insights(result: Dict) -> Dict:
    """Generate insights using Pandas for data analysis"""
    df = pd.DataFrame([result])
    
    insights = {
        "savingsPercentage": (result["savings"] / result["income"]) * 100 if result["income"] > 0 else 0,
        "monthlyImpact": result["monthlySavings"],
        "yearlyImpact": result["savings"],
        "recommendation": result["recommendedRegime"],
        "effectiveRate": result["effectiveRate"]
    }
    
    return insights

@app.get("/")
async def root():
    return {"message": "Taxync API is running", "version": "1.0.0"}

@app.post("/calculate-tax")
async def calculate_tax(request: TaxCalculationRequest):
    """Calculate tax with comprehensive analysis for FY 2025-26"""
    try:
        # Apply Indian FY 2025-26 tax slabs
        income = request.income
        logging.info(
            f"/calculate-tax payload: income={request.income}, section80C={request.section80C}, "
            f"section80D={request.section80D}, hra={request.hra}, home_loan_interest={request.home_loan_interest}, "
            f"standard_deduction={request.standard_deduction}, edu_loan_interest={request.edu_loan_interest}, donations={request.donations}"
        )
        
        # Validate numeric inputs are finite
        for field_name, val in [
            ("income", request.income),
            ("section80C", request.section80C),
            ("section80D", request.section80D),
            ("hra", request.hra),
            ("home_loan_interest", request.home_loan_interest),
            ("standard_deduction", request.standard_deduction),
            ("edu_loan_interest", request.edu_loan_interest),
            ("donations", request.donations),
        ]:
            try:
                if not math.isfinite(float(val)):
                    raise ValueError
            except Exception:
                raise HTTPException(status_code=400, detail=f"Invalid numeric value for {field_name}")
        
        # Create a deductions dictionary for the old regime calculation
        deductions = {
            "section80C": request.section80C,
            "section80D": request.section80D,
            "hra": request.hra,
            "home_loan_interest": request.home_loan_interest,
            "standard_deduction": request.standard_deduction,
            "edu_loan_interest": request.edu_loan_interest,
            "donations": request.donations
        }

        # Use the dedicated calculation functions for consistency and correctness
        total_deductions = sum(deductions.values())
        old_tax = calculate_old_regime_tax(income, total_deductions)
        new_tax = calculate_new_regime_tax(income, request.standard_deduction)

        old_taxable_income = max(0, income - sum(deductions.values()))

        savings = old_tax - new_tax
        logging.info(
            f"/calculate-tax computed: old_taxable_income={old_taxable_income}, new_taxable_income={max(0, income - request.standard_deduction)}, old_tax={old_tax}, new_tax={new_tax}, savings={savings}"
        )
        
        # --- Smart Tax Advisor Logic ---
        suggestions = []
        current_deductions = sum(deductions.values())
        

        # 1. Section 80C Suggestion
        if request.section80C < 150000:
            remaining_80c = 150000 - request.section80C
            potential_saving = get_tax_saving_for_investment(remaining_80c, income, current_deductions)
            if potential_saving > 0:
                suggestions.append(f"Invest ₹{remaining_80c:,.0f} more in Section 80C (e.g., ELSS, PPF) to save up to ₹{potential_saving:,.0f} in taxes.")

        # 2. Section 80D Suggestion (assuming non-senior citizen)
        if request.section80D < 25000:
            remaining_80d = 25000 - request.section80D
            potential_saving = get_tax_saving_for_investment(remaining_80d, income, current_deductions)
            if potential_saving > 0:
                suggestions.append(f"Increase your health insurance premium by ₹{remaining_80d:,.0f} (Section 80D) to save up to ₹{potential_saving:,.0f} in taxes.")

        # 3. Home Loan Interest Suggestion
        if request.home_loan_interest < 200000:
            remaining_interest = 200000 - request.home_loan_interest
            potential_saving = get_tax_saving_for_investment(remaining_interest, income, current_deductions)
            if potential_saving > 0:
                suggestions.append(f"Claiming up to ₹{remaining_interest:,.0f} more in Home Loan Interest (Section 24b) could save you ₹{potential_saving:,.0f}.")

        # 4. NPS Suggestion (Section 80CCD(1B))
        # This is a generic suggestion as we don't have NPS input yet.
        # Only suggest NPS if total income is high enough to benefit.
        if income > 750000: # Threshold where higher tax slabs are hit
            potential_nps_saving = get_tax_saving_for_investment(50000, income, current_deductions)
            if potential_nps_saving > 0:
                suggestions.append(f"Consider investing ₹50,000 in NPS (Section 80CCD(1B)) for an additional tax saving of up to ₹{potential_nps_saving:,.0f}.")

        # --- Format Response ---
        # 1. Recommended Regime Module
        if savings > 0:
            regime_comparison = f"Old Regime saves ₹{abs(savings):,.0f} compared to New Regime."
        elif savings < 0:
            regime_comparison = f"New Regime saves ₹{abs(savings):,.0f} compared to Old Regime."
        else:
            regime_comparison = "Both regimes result in the same tax liability."

        # 2. Tax Optimization Suggestions Module
        if not suggestions:
            suggestions.append("✅ You have already optimized your tax savings under current rules.")
        
        result = {
            "old_regime_tax": round(old_tax),
            "new_regime_tax": round(new_tax),
            "savings": round(savings),
            "regime_comparison": regime_comparison,
            "optimization_suggestions": suggestions,
        }

        # Cache the result in Redis if available
        if redis_client:
            try:
                # Use a hash of the request as the key
                request_dict = request.dict()
                # Sort keys to ensure consistent hashing
                sorted_request_str = json.dumps(request_dict, sort_keys=True)
                cache_key = f"tax_calc:{hashlib.sha256(sorted_request_str.encode()).hexdigest()}"
                redis_client.set(cache_key, json.dumps(result), ex=3600)  # Cache for 1 hour
                logging.info(f"Cached result for key: {cache_key}")
            except Exception as e:
                logging.warning(f"Redis caching failed: {e}")

        return result
    except Exception as e:
        logging.exception("Unhandled error in /calculate-tax")
        raise HTTPException(status_code=500, detail=str(e))

 

@app.post("/export/pdf")
async def export_pdf(request: ExportRequest):
    """Export tax report as PDF with dynamic data and charts"""
    try:
        tax_data = request.taxResult
        form_data = request.formData
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            alignment=TA_CENTER,
            textColor=colors.Color(0.4, 0.2, 0.6)
        )
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            spaceAfter=12,
            textColor=colors.Color(0.4, 0.2, 0.6)
        )
        
        story = []
        story.append(Paragraph("TAXYNC - Tax Analysis Report", title_style))
        story.append(Spacer(1, 20))
        
        # Tax Summary Table
        tax_table_data = [
            ['Description', 'Amount (₹)'],
            ['Annual Income', f"₹{form_data.get('income', 0):,}"],
            ['Old Regime Tax', f"₹{tax_data.get('old_regime_tax', 0):,}"],
            ['New Regime Tax', f"₹{tax_data.get('new_regime_tax', 0):,}"],
            ['Total Savings', f"₹{tax_data.get('savings', 0):,}"]
        ]
        tax_table = Table(tax_table_data, colWidths=[3*inch, 2*inch])
        tax_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.4, 0.2, 0.6)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(tax_table)
        story.append(Spacer(1, 20))

        # Suggestions
        if tax_data.get('suggestions'):
            story.append(Paragraph("<b>Tax Optimization Suggestions:</b>", styles['Heading3']))
            for suggestion in tax_data['suggestions']:
                story.append(Paragraph(f"• {suggestion}", styles['Normal']))
            story.append(Spacer(1, 20))

        # Add chart image if available
        if request.chartImage:
            try:
                # Decode base64 image
                # Expected format: "data:image/png;base64,iVBORw0KGgo..."
                header, encoded = request.chartImage.split(",", 1)
                decoded_image = base64.b64decode(encoded)
                image_buffer = io.BytesIO(decoded_image)
                
                # Add image to PDF
                story.append(Spacer(1, 20))
                story.append(Paragraph("<b>Tax Distribution Chart:</b>", styles['Heading3']))
                img = Image(image_buffer, width=4*inch, height=3*inch)
                img.hAlign = 'CENTER'
                story.append(img)
                story.append(Spacer(1, 20))
            except Exception as e:
                logging.error(f"Failed to process chart image: {e}")

        # Footer
        footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=10, alignment=TA_CENTER, textColor=colors.grey)
        story.append(Spacer(1, 40))
        story.append(Paragraph("Generated by Taxync – Developed by Somil Yadav © 2025", footer_style))
        
        doc.build(story)
        buffer.seek(0)
        
        return StreamingResponse(
            io.BytesIO(buffer.read()),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=Tax_Report_{datetime.now().year}.pdf"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/export/excel")
async def export_excel(request: ExportRequest):
    """Export tax report as Excel with dynamic data and charts"""
    try:
        tax_data = request.taxResult
        form_data = request.formData

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tax Report"

        header_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='663399', end_color='663399', fill_type='solid')
        subheader_font = Font(name='Arial', size=12, bold=True, color='663399')
        center_alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells('A1:E1')
        ws['A1'] = 'TAXYNC - Tax Analysis Report'
        ws['A1'].font = Font(name='Arial', size=18, bold=True, color='663399')
        ws['A1'].alignment = center_alignment

        ws['A3'] = 'Tax Summary'
        ws['A3'].font = subheader_font

        headers = ['Description', 'Amount (₹)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        tax_summary = [
            ['Annual Income', form_data.get('income', 0)],
            ['Old Regime Tax', tax_data.get('old_regime_tax', 0)],
            ['New Regime Tax', tax_data.get('new_regime_tax', 0)],
            ['Total Savings', tax_data.get('savings', 0)]
        ]

        for row, (desc, amount) in enumerate(tax_summary, 5):
            ws.cell(row=row, column=1, value=desc)
            ws.cell(row=row, column=2, value=amount).number_format = '₹#,##0'

        # Detailed Comparison Table
        ws['D3'] = 'Detailed Regime Comparison'
        ws['D3'].font = subheader_font

        comparison_headers = ['Description', 'Old Regime', 'New Regime']
        for col, header in enumerate(comparison_headers, 4):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Safely sum numeric deductions (ignore non-numeric, None)
        def _to_number(x):
            try:
                return float(x)
            except Exception:
                return 0.0

        total_deductions = sum(_to_number(v) for k, v in form_data.items() if k not in ['income', 'name'])
        old_taxable = form_data.get('income', 0) - total_deductions

        comparison_data = [
            ['Gross Income', form_data.get('income', 0), form_data.get('income', 0)],
            ['Deductions', total_deductions, 0],
            ['Taxable Income', old_taxable, form_data.get('income', 0)],
            ['Tax Payable', tax_data.get('old_regime_tax', 0), tax_data.get('new_regime_tax', 0)]
        ]

        for row, data_row in enumerate(comparison_data, 5):
            for col, value in enumerate(data_row, 4):
                cell = ws.cell(row=row, column=col, value=value)
                if row > 5:
                    cell.number_format = '₹#,##0'

        # Suggestions
        if tax_data.get('suggestions'):
            ws['A10'] = 'Tax Optimization Suggestions'
            ws['A10'].font = subheader_font
            for i, suggestion in enumerate(tax_data['suggestions'], 11):
                ws[f'A{i}'] = f"• {suggestion}"

        # Chart (best-effort; don't fail export if chart references are invalid)
        try:
            chart = PieChart()
            chart.title = "Tax Comparison"
            labels = Reference(ws, min_col=1, min_row=6, max_row=8)
            data = Reference(ws, min_col=2, min_row=6, max_row=8)
            chart.add_data(data, titles_from_data=False)
            chart.set_categories(labels)
            ws.add_chart(chart, "D4")
        except Exception as e:
            logging.warning(f"Excel chart generation skipped: {e}")

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[col[0].column_letter].width = adjusted_width

        # Footer
        last_row = ws.max_row + 2
        ws.merge_cells(f'A{last_row}:E{last_row}')
        footer_cell = ws[f'A{last_row}']
        footer_cell.value = "Created with Taxync | Developed by Somil Yadav"
        footer_cell.font = Font(name='Arial', size=10, italic=True, color='808080')
        footer_cell.alignment = center_alignment

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        content = buffer.getvalue()
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=Tax_Report_{datetime.now().year}.xlsx"
            },
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/share")
async def create_shareable_link(request: ShareRequest):
    """Create a shareable link by saving report data to Redis."""
    if not redis_client:
        raise HTTPException(status_code=503, detail="Redis is not available.")
    
    try:
        report_id = str(uuid.uuid4())
        report_data = {
            "formData": request.formData,
            "taxResult": request.taxResult,
            "createdAt": datetime.utcnow().isoformat()
        }
        
        # Store the report data in Redis for 24 hours (86400 seconds)
        redis_client.setex(f"report:{report_id}", 86400, json.dumps(report_data))
        
        return {"reportId": report_id}
    except Exception as e:
        logging.error(f"Failed to create shareable link: {e}")
        raise HTTPException(status_code=500, detail="Could not create shareable link.")

@app.get("/share/{report_id}")
async def get_shared_report(report_id: str):
    """Retrieve a shared report from Redis by its ID."""
    if not redis_client:
        raise HTTPException(status_code=503, detail="Redis is not available.")

    try:
        report_data_json = redis_client.get(f"report:{report_id}")
        if not report_data_json:
            raise HTTPException(status_code=404, detail="Report not found or has expired.")
        
        report_data = json.loads(report_data_json)
        return report_data
    except Exception as e:
        logging.error(f"Failed to retrieve shared report {report_id}: {e}")
        raise HTTPException(status_code=500, detail="Could not retrieve report.")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8001)
